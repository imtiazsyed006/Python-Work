import os.path
from os import path
import os
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl as xl 
names = []
firstRun = []

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(210, 190, 141, 41))
        self.lineEdit.setObjectName("lineEdit")
        font = QtGui.QFont()
        font.setPointSize(14)
        self.lineEdit.setFont(font)
        
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(40, 190, 141, 41))
        
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        
        self.label.setFont(font)
        self.label.setStyleSheet("background-color: rgb(170, 85, 255);")
        self.label.setObjectName("label")
        
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(160, 60, 511, 41))
        self.pushButton.clicked.connect(self.create_excel)
        
        font = QtGui.QFont()
        font.setPointSize(12)
        
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(410, 170, 161, 41))
        
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        
        self.label_2.setFont(font)
        self.label_2.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        
        self.lcdNumber = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber.setGeometry(QtCore.QRect(390, 270, 201, 41))
        self.lcdNumber.setObjectName("lcdNumber")
        
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(390, 220, 201, 41))
        
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(False)
        font.setWeight(50)
        
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("background-color: rgb(224, 185, 255);")
        self.label_3.setObjectName("label_3")
        
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(210, 250, 141, 41))
        self.lineEdit_2.setObjectName("lineEdit_2")
        font = QtGui.QFont()
        font.setPointSize(14)
        self.lineEdit_2.setFont(font)
        
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(40, 250, 141, 41))
        
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        
        self.label_4.setFont(font)
        self.label_4.setStyleSheet("background-color: rgb(170, 85, 255);")
        self.label_4.setObjectName("label_4")
        
        self.lineEdit_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_3.setGeometry(QtCore.QRect(210, 310, 141, 41))
        self.lineEdit_3.setObjectName("lineEdit_3")
        font = QtGui.QFont()
        font.setPointSize(14)
        self.lineEdit_3.setFont(font)
        
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(40, 310, 141, 41))
        
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("background-color: rgb(170, 85, 255);")
        self.label_5.setObjectName("label_5")
        
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(210, 380, 141, 41))
        self.pushButton_2.clicked.connect(self.upload) #############
        
        font = QtGui.QFont()
        font.setPointSize(12)
        
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(390, 320, 201, 41))
        self.pushButton_3.clicked.connect(self.open_excel)
        
        font = QtGui.QFont()
        font.setPointSize(12)
        
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")
        
        self.lineEdit_4 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_4.setGeometry(QtCore.QRect(160, 10, 511, 41))
        self.lineEdit_4.setObjectName("lineEdit_4")
        font = QtGui.QFont()
        font.setPointSize(14)
        self.lineEdit_4.setFont(font)
        
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(140, 490, 451, 41))
        
        font = QtGui.QFont()
        font.setPointSize(17)

        self.plainTextEdit = QtWidgets.QPlainTextEdit(self.centralwidget)
        self.plainTextEdit.setGeometry(QtCore.QRect(610, 210, 181, 341))
        self.plainTextEdit.setObjectName("plainTextEdit")
        self.plainTextEdit.setReadOnly(True)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.plainTextEdit.setFont(font)
        
        self.label_6.setFont(font)
        self.label_6.setStyleSheet("background-color: rgb(228, 211, 255);")
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(50, 490, 91, 41))
        
        font = QtGui.QFont()
        font.setPointSize(15)
        
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("background-color: rgb(230, 212, 255);")
        self.label_7.setObjectName("label_7")
        
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(10, 10, 141, 41))
        
        font = QtGui.QFont()
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("background-color: rgb(170, 85, 255);")
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "Reg Number"))
        self.pushButton.setText(_translate("MainWindow", "Create Sheet"))
        self.label_2.setText(_translate("MainWindow", "Hit enter to upload"))
        self.label_3.setText(_translate("MainWindow", "Total number of entries"))
        self.label_4.setText(_translate("MainWindow", "Quiz"))
        self.label_5.setText(_translate("MainWindow", "Assignment"))
        self.pushButton_2.setText(_translate("MainWindow", "Upload"))
        self.pushButton_3.setText(_translate("MainWindow", "Open Excel File"))
        self.label_6.setText(_translate("MainWindow", ""))
        self.label_7.setText(_translate("MainWindow", "Status"))
        self.label_8.setText(_translate("MainWindow", "Sheet Name"))

    def create_excel(self):
        global file_Name
        file_Name = self.lineEdit_4.text()
        names.append(file_Name)
        self.plainTextEdit.clear()
        if not file_Name:
            self.plainTextEdit.insertPlainText('Enter a valid file name')
        elif path.exists(str(file_Name)+'.xlsx') == 1: 
            self.plainTextEdit.insertPlainText('Go ahead file exists already')      
        else:
            title = ['Reg No','Quiz','Assignment']
            df = pd.DataFrame(columns = title)
            writer = pd.ExcelWriter(str(file_Name)+'.xlsx', engine = 'xlsxwriter')
            df.to_excel(writer,sheet_name = 'Sheet1',header = True,index = False)
            writer.save()
            self.plainTextEdit.insertPlainText('Excel file created')
        self.lineEdit_4.clear()

    def upload(self):
        regNo = self.lineEdit.text()
        quiz  = self.lineEdit_2.text()
        assignment = self.lineEdit_3.text()
        name = self.lineEdit_4.text()
        self.plainTextEdit.clear()

        if not file_Name:
            self.plainTextEdit.insertPlainText('Create an Excel file first')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) == 0 and len(str(quiz)) != 0 and len(str(assignment)) != 0:
            self.plainTextEdit.insertPlainText('Enter Reg No')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) != 0 and len(str(quiz)) == 0 and len(str(assignment)) != 0:
            self.plainTextEdit.insertPlainText('Enter quiz marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) != 0 and len(str(quiz)) != 0 and len(str(assignment)) == 0:
            self.plainTextEdit.insertPlainText('Enter assignment marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) == 0 and len(str(quiz)) == 0 and len(str(assignment)) != 0:
            self.plainTextEdit.insertPlainText('Enter regNo and quiz marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) == 0 and len(str(quiz)) != 0 and len(str(assignment)) == 0:
            self.plainTextEdit.insertPlainText('Enter regNo and assignment marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) != 0 and len(str(quiz)) == 0 and len(str(assignment)) == 0:
            self.plainTextEdit.insertPlainText('Enter quiz and assignment marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        elif not not names and len(str(regNo)) == 0 and len(str(quiz)) ==0  and len(str(assignment)) == 0:
            print(name)
            self.plainTextEdit.insertPlainText('Enter Reg No, Quiz and Assignment marks')
            self.label_6.setText('Some thing is wrong')
            self.label_6.setStyleSheet("background-color: rgb(255, 0, 0);")

        else:
            self.plainTextEdit.insertPlainText('Uploaded Successfully')
            self.label_6.setText('Everything is fine')
            self.label_6.setStyleSheet("background-color: rgb(224, 185, 255);")
            wb = xl.load_workbook(names[-1]+'.xlsx',read_only = False)
            ws = wb.active

            noRows_1 = ws.max_row
            for i in range(1,noRows_1+2):
                if ws.cell(i,1).value == int(regNo):
                    c_1 = ws.cell(i,1)
                    c_2 = ws.cell(i,2)
                    c_3 = ws.cell(i,3)
                    c_1.value = int(regNo)
                    c_2.value = int(quiz)
                    c_3.value = int(assignment)
                    wb.save(names[-1]+'.xlsx')
                    break
                elif i == noRows_1+1:
                    c_1 = ws.cell(noRows_1+1,1)
                    c_2 = ws.cell(noRows_1+1,2)
                    c_3 = ws.cell(noRows_1+1,3)
                    c_1.value = int(regNo)
                    c_2.value = int(quiz)
                    c_3.value = int(assignment)
                    wb.save(names[-1]+'.xlsx')

            freshData = pd.read_excel(names[-1]+'.xlsx',sheet_name = 'Sheet1')
            freshData = freshData.sort_values(by = 'Reg No')  
            freshData.to_excel(names[-1]+'.xlsx',index = False, header = True)
            print(freshData)  
            self.lcdNumber.display(noRows_1)
            self.lineEdit.clear()
            self.lineEdit_2.clear()
            self.lineEdit_3.clear()    


    def open_excel(self):
        os.startfile(names[-1]+".xlsx")

    def get_lineEdit(self):
        regNo = self.lineEdit.text()
        return regNo
    
    def get_lineEdit_4(self):
        global fileName
        fileName = self.lineEdit_4.text()
        fileName = str(fileName)
        return (fileName)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
